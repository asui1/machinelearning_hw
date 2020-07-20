'''
Created on 2017. 11. 6.

@author: Asui
'''
import numpy as np
import random
import copy
from numpy.linalg import inv
import math
from openpyxl import Workbook

#variables
#val : [[y, x1, x2 .... x13], ...]
vals = []
ten_random = []
cut = 0
ys = []
xs = []
xs2 = []
tempxs = []
tempys = []
tempxt = []
tempyt = []


def generate_random(list_len):
    initial_list = []
    for i in range(list_len):
        initial_list.append(i)
    for i in range(10):
        random.shuffle(initial_list)
        new_list = copy.copy(initial_list)
        ten_random.append(new_list)

def read_data():
    f = open("E:ml_1_data.txt", 'r')
    lines = f.readlines()
    for line in lines:
        this_line = line.replace(":", " ")
        this_line = this_line.split()
        vals.append([])
        xs.append([])
        xs2.append([])
        for i in range(0, len(this_line), 2):
            vals[len(vals)-1].append(this_line[i])
            if i != 0:
                xs[len(vals)-1].append(this_line[i])
                xs2[len(vals)-1].append(this_line[i])
            else:
                ys.append(this_line[0])
        xs[len(vals)-1].append(1)
    f.close()

#remove colons in text
def remove_col(alist):
    for i in range(len(alist)):
        alist[i] = alist[i].parse
        
def get_error(real_ys, x_array, beta):
    total_error = 0
    temp_y = 0
    for i in range(len(real_ys)):
        temp_y = np.dot(np.transpose(beta), x_array[i])
        total_error += abs(temp_y-real_ys[i])
    return (total_error/len(real_ys))

def data_error(real_ys, x_array, beta):
    total_error = 0
    temp_y = 0
    for i in range(len(real_ys)):
        temp_y = np.dot(np.transpose(beta), x_array[i])
        total_error += abs(temp_y-real_ys[i])
    return ("dataerror:"+str(total_error/len(real_ys)))

read_data()
generate_random(len(vals))


wb1 = Workbook()
ws1 = wb1.active

def data_randomx(x, random_array):
    tempxs = []
    for i in range(404):
        tempxs.append(x[random_array[i]])
    return tempxs 

def data_randomy(y, random_array):
    tempys = []
    for i in range(404):
        tempys.append(y[random_array[i]])
    return tempys
        
def data_testx(x, random_array):
    tempxt = []
    for i in range(404, len(random_array)):
        tempxt.append(x[random_array[i]])
    return tempxt

def data_testy(y, random_array):
    tempyt = []
    for i in range(404, len(random_array)):
        tempyt.append(y[random_array[i]])
    return tempyt
                
def prob1(x_array, y_array, random_array):
    tempxs = data_randomx(xs, random_array)
    tempys = data_randomy(ys, random_array)
    tempxt = data_testx(xs, random_array)
    tempyt = data_testy(ys, random_array)
    large_x = np.array(tempxs, dtype=float)
    large_y = np.array(tempys, dtype=float)
    large_xt = np.array(tempxt, dtype=float)
    large_yt = np.array(tempyt, dtype=float)
    trans_x = np.transpose(large_x)
    tempmult = inv(np.dot(trans_x, large_x))
    beta1 = np.dot(np.dot(tempmult, trans_x), large_y)
    return get_error(large_yt, large_xt, beta1)

def prob1_2(x_array, y_array, random_array):
    tempxs = data_randomx(xs2, random_array)
    tempys = data_randomy(ys, random_array)
    tempxt = data_testx(xs2, random_array)
    tempyt = data_testy(ys, random_array)
    large_x = np.array(tempxs, dtype=float)
    large_y = np.array(tempys, dtype=float)
    large_xt = np.array(tempxt, dtype=float)
    large_yt = np.array(tempyt, dtype=float)
    trans_x = np.transpose(large_x)
    tempmult = inv(np.dot(trans_x, large_x))
    beta1 = np.dot(np.dot(tempmult, trans_x), large_y)
    return get_error(large_yt, large_xt, beta1)
    
for i in range(10):
    ws1['A'+str(i+1)] = prob1(xs, ys, ten_random[i])
    

for i in range(10):
    ws1['C'+str(i+1)] = prob1_2(xs2, ys, ten_random[i])
    
wb1.save("ML_Homework1_1.xlsx")
print("prob1 fin")    


wb2 = Workbook()
ws2 = wb2.active


def gradient(beta_vector, y_vector, x_vector):
    total = np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float)
    for i in range(len(y_vector)):
        temp_total = y_vector[i] - np.asscalar(np.dot(np.transpose(beta_vector), x_vector[i]))
        total += temp_total * x_vector[i]
    return total * 2 / len(y_vector)
    
        

def prob2(beta, step_size, num_iter, random_array, sheet, letter):
    tempxs = data_randomx(xs, random_array)
    tempys = data_randomy(ys, random_array)
    tempxt = data_testx(xs, random_array)
    tempyt = data_testy(ys, random_array)
    large_x = np.array(tempxs, dtype=float)
    large_y = np.array(tempys, dtype=float)
    large_xt = np.array(tempxt, dtype=float)
    large_yt = np.array(tempyt, dtype=float)
    for i in range(num_iter):
        beta += step_size * gradient(beta, large_y, large_x)
        sheet[letter+str(i+1)]=get_error(large_yt, large_xt, beta)
        #data_error(large_y, large_x, beta)
    return get_error(large_yt, large_xt, beta)

for i in range(10):
    ws2['C'+str(i+1)] = prob2(np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float), 1.0, 200, ten_random[i], ws2, 'A')
    ws2['G'+str(i+1)] = prob2(np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float), 0.05, 200, ten_random[i], ws2, 'E')
    ws2['K'+str(i+1)] = prob2(np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float), 0.0001, 200, ten_random[i], ws2, 'I')
   
wb2.save("ML_Homework1_2.xlsx")
print("prob2 fin")    

wb3 = Workbook()
ws3 = wb3.active

def coord(beta_vector, y_vector, x_vector, n):
    total = np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float)
    for i in range(len(y_vector)):
        temp_total = y_vector[i] - np.asscalar(np.dot(np.transpose(beta_vector), x_vector[i]))
        total += temp_total * x_vector[i][n]
    return total * 2 / len(y_vector)

def prob3(beta, step_size, num_iter, random_array, sheet):
    tempxs = data_randomx(xs, random_array)
    tempys = data_randomy(ys, random_array)
    tempxt = data_testx(xs, random_array)
    tempyt = data_testy(ys, random_array)
    large_x = np.array(tempxs, dtype=float)
    large_y = np.array(tempys, dtype=float)
    large_xt = np.array(tempxt, dtype=float)
    large_yt = np.array(tempyt, dtype=float)
    for j in range(len(beta)):
        for i in range(num_iter):
            beta += step_size * coord(beta, large_y, large_x, j)
            sheet['I'+str(i*len(beta)+j+1)] = get_error(large_yt, large_xt, beta)
            #data_error(large_y, large_x, beta)
    return get_error(large_yt, large_xt, beta)

for i in range(10):
    ws3['K'+str(i+1)] = prob3(np.array([0,0,0,0,0,0,0,0,0,0,0,0,0,0], dtype=float), 0.05, 500, ten_random[i], ws3)

wb3.save("ML_Homework1_3.xlsx")
print("prob3 fin")    


